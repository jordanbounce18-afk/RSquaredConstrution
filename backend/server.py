from fastapi import FastAPI, APIRouter, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Email integration (Emergent-managed Resend proxy)
EMAIL_BASE_URL = "https://integrations.emergentagent.com"
EMAIL_KEY = os.environ.get("EMERGENT_EMAIL_KEY")
EMAIL_FROM_NAME = os.environ.get("EMAIL_FROM_NAME", "R2 Construction")
ESTIMATE_NOTIFY_EMAIL = os.environ.get("ESTIMATE_NOTIFY_EMAIL")

app = FastAPI(title="R2 Construction API")
api_router = APIRouter(prefix="/api")


# ============ Models ============
class EstimateRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))

    # Basic Information
    name: str
    email: EmailStr
    phone: str
    preferred_contact: Optional[str] = None  # Email / Phone / Either
    best_time_to_contact: Optional[str] = None  # Morning / Afternoon / Evening / Anytime
    address: Optional[str] = None
    property_type: Optional[str] = None  # Single-family, Townhome, Condo, Other

    # Project Details
    project_type: str  # legacy primary type
    project_types: Optional[List[str]] = None  # multi-select
    scope: Optional[str] = None
    square_footage: Optional[str] = None
    style_preference: Optional[str] = None
    has_plans: Optional[str] = None  # Yes / No / In progress

    # Timeline
    ideal_start_date: Optional[str] = None  # ISO date
    timeline: Optional[str] = None
    hard_deadline: Optional[str] = None

    # Budget
    budget: Optional[str] = None
    financing: Optional[str] = None
    budget_flexibility: Optional[str] = None

    # Message + meta
    message: str
    hear_about: Optional[str] = None

    status: str = "new"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class EstimateCreate(BaseModel):
    # Basic
    name: str = Field(..., min_length=1, max_length=120)
    email: EmailStr
    phone: str = Field(..., min_length=5, max_length=40)
    preferred_contact: Optional[str] = None
    best_time_to_contact: Optional[str] = None
    address: Optional[str] = None
    property_type: Optional[str] = None

    # Project
    project_type: str = Field(..., min_length=1)
    project_types: Optional[List[str]] = None
    scope: Optional[str] = None
    square_footage: Optional[str] = None
    style_preference: Optional[str] = None
    has_plans: Optional[str] = None

    # Timeline
    ideal_start_date: Optional[str] = None
    timeline: Optional[str] = None
    hard_deadline: Optional[str] = None

    # Budget
    budget: Optional[str] = None
    financing: Optional[str] = None
    budget_flexibility: Optional[str] = None

    # Message + meta
    message: str = Field(..., min_length=1, max_length=6000)
    hear_about: Optional[str] = None


class EstimateStatusUpdate(BaseModel):
    status: str


# ============ Email helpers ============
FIELD_LABELS = [
    ("BASIC INFORMATION", None),
    ("Name", "name"),
    ("Email", "email"),
    ("Phone", "phone"),
    ("Preferred Contact", "preferred_contact"),
    ("Best Time to Contact", "best_time_to_contact"),
    ("Property Address", "address"),
    ("Property Type", "property_type"),
    ("PROJECT DETAILS", None),
    ("Primary Project Type", "project_type"),
    ("Additional Project Types", "project_types"),
    ("Scope of Work", "scope"),
    ("Approx. Square Footage", "square_footage"),
    ("Style Preference", "style_preference"),
    ("Has Plans / Drawings?", "has_plans"),
    ("TIMELINE", None),
    ("Ideal Start Date", "ideal_start_date"),
    ("Timeline Flexibility", "timeline"),
    ("Hard Deadlines / Events", "hard_deadline"),
    ("BUDGET", None),
    ("Budget Range", "budget"),
    ("Financing Approach", "financing"),
    ("Budget Flexibility", "budget_flexibility"),
    ("MESSAGE & SOURCE", None),
    ("Project Message / Details", "message"),
    ("How they found us", "hear_about"),
]


def _fmt_value(v):
    if v is None or v == "":
        return "—"
    if isinstance(v, list):
        return ", ".join(v) if v else "—"
    return str(v)


def build_estimate_email_html(est: EstimateRequest) -> str:
    d = est.model_dump()
    created = est.created_at.strftime("%B %d, %Y — %I:%M %p UTC")

    rows_html = []
    for label, key in FIELD_LABELS:
        if key is None:
            rows_html.append(
                f'<tr><td colspan="2" style="background:#1C1C1C;color:#FAF9F6;'
                f'padding:12px 16px;font-size:11px;letter-spacing:2px;'
                f'text-transform:uppercase;font-family:Arial,sans-serif;">'
                f'{label}</td></tr>'
            )
        else:
            val = _fmt_value(d.get(key))
            rows_html.append(
                f'<tr>'
                f'<td style="width:35%;padding:10px 16px;border-bottom:1px solid #DCD7CE;'
                f'background:#F2EFE9;font-family:Arial,sans-serif;font-size:12px;'
                f'color:#595959;vertical-align:top;">{label}</td>'
                f'<td style="padding:10px 16px;border-bottom:1px solid #DCD7CE;'
                f'font-family:Arial,sans-serif;font-size:13px;color:#1C1C1C;'
                f'vertical-align:top;white-space:pre-wrap;">{val}</td>'
                f'</tr>'
            )

    return f"""
<!DOCTYPE html>
<html><body style="margin:0;padding:0;background:#FAF9F6;">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#FAF9F6;">
    <tr><td align="center" style="padding:32px 12px;">
      <table role="presentation" width="640" cellspacing="0" cellpadding="0"
             style="max-width:640px;background:#FAF9F6;border:1px solid #DCD7CE;">
        <tr>
          <td style="padding:32px 32px 8px 32px;font-family:Georgia,serif;">
            <div style="font-size:11px;letter-spacing:3px;color:#9E907F;text-transform:uppercase;">
              R² Construction — New Estimate Request
            </div>
            <h1 style="margin:12px 0 4px 0;font-size:30px;font-weight:400;color:#1C1C1C;">
              {est.name}
            </h1>
            <div style="font-size:13px;color:#595959;font-family:Arial,sans-serif;">
              Received {created} · ID {est.id}
            </div>
          </td>
        </tr>
        <tr><td style="padding:16px 32px 32px 32px;">
          <table role="presentation" width="100%" cellspacing="0" cellpadding="0"
                 style="border:1px solid #DCD7CE;border-collapse:collapse;">
            {''.join(rows_html)}
          </table>
          <p style="margin:24px 0 0 0;font-family:Arial,sans-serif;font-size:12px;
                    color:#595959;line-height:1.6;">
            Reply directly to this email to reach {est.name} at
            <a href="mailto:{est.email}" style="color:#1C1C1C;">{est.email}</a>
            or call <a href="tel:{est.phone}" style="color:#1C1C1C;">{est.phone}</a>.
          </p>
        </td></tr>
        <tr><td style="padding:16px 32px;background:#1C1C1C;color:#E8E4DB;
                      font-family:Arial,sans-serif;font-size:11px;letter-spacing:2px;
                      text-transform:uppercase;">
          R² Construction · Automated notification
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>
"""


async def send_estimate_email(est: EstimateRequest) -> None:
    if not EMAIL_KEY or not ESTIMATE_NOTIFY_EMAIL:
        logger.warning("Email notification skipped: EMAIL_KEY or ESTIMATE_NOTIFY_EMAIL not set")
        return
    payload = {
        "to": [ESTIMATE_NOTIFY_EMAIL],
        "subject": f"New Estimate Request — {est.name} ({est.project_type})",
        "html": build_estimate_email_html(est),
        "from_name": EMAIL_FROM_NAME,
        "contact_email": est.email,
    }
    try:
        async with httpx.AsyncClient(timeout=30) as http:
            resp = await http.post(
                f"{EMAIL_BASE_URL}/api/v1/email/send",
                headers={"X-Email-Key": EMAIL_KEY},
                json=payload,
            )
        if resp.status_code >= 300:
            logger.error(f"Email send failed: {resp.status_code} {resp.text}")
        else:
            logger.info(f"Estimate email sent to {ESTIMATE_NOTIFY_EMAIL} (id={est.id})")
    except Exception as e:
        logger.exception(f"Email send error: {e}")


# ============ Excel export ============
EXPORT_HEADERS = [
    ("Received", "created_at"),
    ("Status", "status"),
    ("Name", "name"),
    ("Email", "email"),
    ("Phone", "phone"),
    ("Preferred Contact", "preferred_contact"),
    ("Best Time", "best_time_to_contact"),
    ("Address", "address"),
    ("Property Type", "property_type"),
    ("Primary Project", "project_type"),
    ("Other Project Types", "project_types"),
    ("Scope", "scope"),
    ("Sq. Ft.", "square_footage"),
    ("Style", "style_preference"),
    ("Has Plans", "has_plans"),
    ("Ideal Start", "ideal_start_date"),
    ("Timeline", "timeline"),
    ("Hard Deadline", "hard_deadline"),
    ("Budget", "budget"),
    ("Financing", "financing"),
    ("Budget Flexibility", "budget_flexibility"),
    ("Heard About Us", "hear_about"),
    ("Message", "message"),
    ("ID", "id"),
]


def build_estimates_workbook(items: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimates"

    header_font = Font(name="Calibri", bold=True, color="FAF9F6", size=11)
    header_fill = PatternFill("solid", fgColor="1C1C1C")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col_idx, (label, _key) in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, it in enumerate(items, start=2):
        for col_idx, (_label, key) in enumerate(EXPORT_HEADERS, start=1):
            v = it.get(key)
            if isinstance(v, list):
                v = ", ".join(v)
            if key == "created_at" and v:
                try:
                    dt = v if isinstance(v, datetime) else datetime.fromisoformat(str(v))
                    v = dt.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    v = str(v)
            cell = ws.cell(row=row_idx, column=col_idx, value=v if v is not None else "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [18, 12, 22, 28, 18, 16, 14, 32, 16, 22, 26, 40, 12, 18, 14, 14, 16, 22, 18, 22, 18, 22, 60, 40]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============ Routes ============
@api_router.get("/")
async def root():
    return {"message": "R2 Construction API", "status": "ok"}


@api_router.post("/estimates", response_model=EstimateRequest, status_code=201)
async def create_estimate(payload: EstimateCreate, background: BackgroundTasks):
    estimate = EstimateRequest(**payload.model_dump())
    doc = estimate.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.estimates.insert_one(doc)
    background.add_task(send_estimate_email, estimate)
    return estimate


@api_router.get("/estimates", response_model=List[EstimateRequest])
async def list_estimates():
    cursor = db.estimates.find({}, {"_id": 0}).sort("created_at", -1)
    items = await cursor.to_list(1000)
    for it in items:
        if isinstance(it.get('created_at'), str):
            it['created_at'] = datetime.fromisoformat(it['created_at'])
    return items


@api_router.get("/estimates/export.xlsx")
async def export_estimates_xlsx():
    cursor = db.estimates.find({}, {"_id": 0}).sort("created_at", -1)
    items = await cursor.to_list(5000)
    data = build_estimates_workbook(items)
    filename = f"r2-estimates-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/estimates/stats/summary")
async def estimates_summary():
    total = await db.estimates.count_documents({})
    new = await db.estimates.count_documents({"status": "new"})
    reviewed = await db.estimates.count_documents({"status": "reviewed"})
    contacted = await db.estimates.count_documents({"status": "contacted"})
    closed = await db.estimates.count_documents({"status": "closed"})
    return {"total": total, "new": new, "reviewed": reviewed,
            "contacted": contacted, "closed": closed}


@api_router.get("/estimates/{estimate_id}", response_model=EstimateRequest)
async def get_estimate(estimate_id: str):
    doc = await db.estimates.find_one({"id": estimate_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Estimate not found")
    if isinstance(doc.get('created_at'), str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    return doc


@api_router.patch("/estimates/{estimate_id}", response_model=EstimateRequest)
async def update_estimate_status(estimate_id: str, payload: EstimateStatusUpdate):
    allowed = {"new", "reviewed", "contacted", "closed"}
    if payload.status not in allowed:
        raise HTTPException(status_code=400, detail=f"Invalid status. Allowed: {sorted(allowed)}")
    result = await db.estimates.update_one(
        {"id": estimate_id}, {"$set": {"status": payload.status}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Estimate not found")
    doc = await db.estimates.find_one({"id": estimate_id}, {"_id": 0})
    if isinstance(doc.get('created_at'), str):
        doc['created_at'] = datetime.fromisoformat(doc['created_at'])
    return doc


@api_router.delete("/estimates/{estimate_id}", status_code=204)
async def delete_estimate(estimate_id: str):
    result = await db.estimates.delete_one({"id": estimate_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Estimate not found")
    return None


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
